"""
Storage and session management utilities.
Handles logging, results storage, and historical tracking.
"""

import json
import pickle
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union
import logging
import os
from datetime import datetime, timedelta
import uuid
from pathlib import Path

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class SessionManager:
    """Manages user sessions and test runs."""
    
    def __init__(self, storage_dir: str = "sessions"):
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(exist_ok=True)
        self.current_session = None
        self.sessions = {}
        
        # Load existing sessions
        self._load_sessions()
    
    def create_session(self, session_name: Optional[str] = None) -> str:
        """
        Create a new session.
        
        Args:
            session_name: Optional name for the session
            
        Returns:
            str: Session ID
        """
        session_id = str(uuid.uuid4())
        
        if session_name is None:
            session_name = f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        session_data = {
            'session_id': session_id,
            'session_name': session_name,
            'created_at': datetime.now().isoformat(),
            'last_updated': datetime.now().isoformat(),
            'test_runs': [],
            'model_info': {},
            'schema_info': {},
            'status': 'active'
        }
        
        self.sessions[session_id] = session_data
        self.current_session = session_id
        
        # Save session
        self._save_session(session_id)
        
        logger.info(f"Created new session: {session_name} ({session_id})")
        return session_id
    
    def load_session(self, session_id: str) -> bool:
        """
        Load an existing session.
        
        Args:
            session_id: Session ID to load
            
        Returns:
            bool: True if successful
        """
        if session_id in self.sessions:
            self.current_session = session_id
            logger.info(f"Loaded session: {session_id}")
            return True
        else:
            logger.warning(f"Session not found: {session_id}")
            return False
    
    def get_current_session(self) -> Optional[Dict[str, Any]]:
        """Get current session data."""
        if self.current_session:
            return self.sessions.get(self.current_session)
        return None
    
    def add_test_run(self, test_results: Dict[str, Any], 
                    test_name: Optional[str] = None) -> str:
        """
        Add a test run to the current session.
        
        Args:
            test_results: Results from the test
            test_name: Optional name for the test
            
        Returns:
            str: Test run ID
        """
        if not self.current_session:
            raise ValueError("No active session. Create a session first.")
        
        test_run_id = str(uuid.uuid4())
        
        if test_name is None:
            test_name = f"test_{datetime.now().strftime('%H%M%S')}"
        
        test_run = {
            'test_run_id': test_run_id,
            'test_name': test_name,
            'timestamp': datetime.now().isoformat(),
            'results': test_results,
            'status': 'completed'
        }
        
        # Add to current session
        self.sessions[self.current_session]['test_runs'].append(test_run)
        self.sessions[self.current_session]['last_updated'] = datetime.now().isoformat()
        
        # Save session
        self._save_session(self.current_session)
        
        logger.info(f"Added test run: {test_name} ({test_run_id})")
        return test_run_id
    
    def update_model_info(self, model_info: Dict[str, Any]):
        """Update model information for the current session."""
        if not self.current_session:
            raise ValueError("No active session. Create a session first.")
        
        self.sessions[self.current_session]['model_info'] = model_info
        self.sessions[self.current_session]['last_updated'] = datetime.now().isoformat()
        
        # Save session
        self._save_session(self.current_session)
    
    def update_schema_info(self, schema_info: Dict[str, Any]):
        """Update schema information for the current session."""
        if not self.current_session:
            raise ValueError("No active session. Create a session first.")
        
        self.sessions[self.current_session]['schema_info'] = schema_info
        self.sessions[self.current_session]['last_updated'] = datetime.now().isoformat()
        
        # Save session
        self._save_session(self.current_session)
    
    def get_test_runs(self, session_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get test runs for a session."""
        if session_id is None:
            session_id = self.current_session
        
        if session_id and session_id in self.sessions:
            return self.sessions[session_id]['test_runs']
        
        return []
    
    def get_session_summary(self, session_id: Optional[str] = None) -> Dict[str, Any]:
        """Get summary of a session."""
        if session_id is None:
            session_id = self.current_session
        
        if not session_id or session_id not in self.sessions:
            return {'error': 'Session not found'}
        
        session = self.sessions[session_id]
        test_runs = session['test_runs']
        
        summary = {
            'session_id': session_id,
            'session_name': session['session_name'],
            'created_at': session['created_at'],
            'last_updated': session['last_updated'],
            'total_test_runs': len(test_runs),
            'model_info': session['model_info'],
            'schema_info': session['schema_info'],
            'status': session['status']
        }
        
        # Add performance trends if available
        if test_runs:
            performance_trends = self._calculate_performance_trends(test_runs)
            summary['performance_trends'] = performance_trends
        
        return summary
    
    def _calculate_performance_trends(self, test_runs: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate performance trends from test runs."""
        trends = {}
        
        # Extract performance metrics
        latencies = []
        throughputs = []
        accuracies = []
        
        for test_run in test_runs:
            performance = test_run.get('results', {}).get('performance', {})
            
            if 'latency_mean_ms' in performance:
                latencies.append(performance['latency_mean_ms'])
            
            if 'throughput_preds_per_sec' in performance:
                throughputs.append(performance['throughput_preds_per_sec'])
            
            if 'accuracy' in performance and performance['accuracy'] is not None:
                accuracies.append(performance['accuracy'])
        
        # Calculate trends
        if len(latencies) > 1:
            trends['latency_trend'] = 'improving' if latencies[-1] < latencies[0] else 'degrading'
            trends['latency_change_percent'] = ((latencies[-1] - latencies[0]) / latencies[0] * 100) if latencies[0] > 0 else 0
        
        if len(throughputs) > 1:
            trends['throughput_trend'] = 'improving' if throughputs[-1] > throughputs[0] else 'degrading'
            trends['throughput_change_percent'] = ((throughputs[-1] - throughputs[0]) / throughputs[0] * 100) if throughputs[0] > 0 else 0
        
        if len(accuracies) > 1:
            trends['accuracy_trend'] = 'improving' if accuracies[-1] > accuracies[0] else 'degrading'
            trends['accuracy_change_percent'] = ((accuracies[-1] - accuracies[0]) / accuracies[0] * 100) if accuracies[0] > 0 else 0
        
        return trends
    
    def list_sessions(self) -> List[Dict[str, Any]]:
        """List all available sessions."""
        sessions_list = []
        
        for session_id, session_data in self.sessions.items():
            sessions_list.append({
                'session_id': session_id,
                'session_name': session_data['session_name'],
                'created_at': session_data['created_at'],
                'last_updated': session_data['last_updated'],
                'total_test_runs': len(session_data['test_runs']),
                'status': session_data['status']
            })
        
        # Sort by last updated
        sessions_list.sort(key=lambda x: x['last_updated'], reverse=True)
        
        return sessions_list
    
    def delete_session(self, session_id: str) -> bool:
        """Delete a session."""
        if session_id in self.sessions:
            del self.sessions[session_id]
            
            # Delete file
            session_file = self.storage_dir / f"{session_id}.json"
            if session_file.exists():
                session_file.unlink()
            
            # Clear current session if it was deleted
            if self.current_session == session_id:
                self.current_session = None
            
            logger.info(f"Deleted session: {session_id}")
            return True
        
        return False
    
    def export_session(self, session_id: str, export_path: str) -> bool:
        """Export session data to a file."""
        if session_id not in self.sessions:
            return False
        
        try:
            session_data = self.sessions[session_id]
            
            with open(export_path, 'w') as f:
                json.dump(session_data, f, indent=2, default=str)
            
            logger.info(f"Exported session {session_id} to {export_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export session: {str(e)}")
            return False
    
    def _save_session(self, session_id: str):
        """Save session data to file."""
        if session_id not in self.sessions:
            return
        
        session_file = self.storage_dir / f"{session_id}.json"
        
        try:
            with open(session_file, 'w') as f:
                json.dump(self.sessions[session_id], f, indent=2, default=str)
        except Exception as e:
            logger.error(f"Failed to save session {session_id}: {str(e)}")
    
    def _load_sessions(self):
        """Load existing sessions from storage directory."""
        if not self.storage_dir.exists():
            return
        
        for session_file in self.storage_dir.glob("*.json"):
            try:
                with open(session_file, 'r') as f:
                    session_data = json.load(f)
                
                session_id = session_data.get('session_id')
                if session_id:
                    self.sessions[session_id] = session_data
                    
            except Exception as e:
                logger.warning(f"Failed to load session from {session_file}: {str(e)}")


class ResultsStorage:
    """Handles storage of test results and visualizations."""
    
    def __init__(self, storage_dir: str = "results"):
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(exist_ok=True)
    
    def save_results(self, results: Dict[str, Any], 
                    filename: Optional[str] = None) -> str:
        """
        Save test results to file.
        
        Args:
            results: Results dictionary to save
            filename: Optional filename (auto-generated if None)
            
        Returns:
            str: Path to saved file
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"results_{timestamp}.json"
        
        file_path = self.storage_dir / filename
        
        try:
            with open(file_path, 'w') as f:
                json.dump(results, f, indent=2, default=str)
            
            logger.info(f"Saved results to {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to save results: {str(e)}")
            raise
    
    def load_results(self, filename: str) -> Dict[str, Any]:
        """Load results from file."""
        file_path = self.storage_dir / filename
        
        try:
            with open(file_path, 'r') as f:
                results = json.load(f)
            
            logger.info(f"Loaded results from {file_path}")
            return results
            
        except Exception as e:
            logger.error(f"Failed to load results: {str(e)}")
            raise
    
    def save_visualization(self, image_data: str, 
                          filename: Optional[str] = None) -> str:
        """
        Save visualization image.
        
        Args:
            image_data: Base64 encoded image data
            filename: Optional filename
            
        Returns:
            str: Path to saved file
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"viz_{timestamp}.png"
        
        file_path = self.storage_dir / filename
        
        try:
            import base64
            image_bytes = base64.b64decode(image_data)
            
            with open(file_path, 'wb') as f:
                f.write(image_bytes)
            
            logger.info(f"Saved visualization to {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to save visualization: {str(e)}")
            raise
    
    def list_results(self) -> List[Dict[str, Any]]:
        """List all saved results files."""
        results_list = []
        
        for file_path in self.storage_dir.glob("*.json"):
            try:
                stat = file_path.stat()
                results_list.append({
                    'filename': file_path.name,
                    'size': stat.st_size,
                    'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
            except Exception as e:
                logger.warning(f"Failed to get info for {file_path}: {str(e)}")
        
        # Sort by modification time
        results_list.sort(key=lambda x: x['modified'], reverse=True)
        
        return results_list
    
    def cleanup_old_results(self, days_old: int = 30):
        """Clean up results older than specified days."""
        cutoff_date = datetime.now() - timedelta(days=days_old)
        deleted_count = 0
        
        for file_path in self.storage_dir.glob("*"):
            try:
                if datetime.fromtimestamp(file_path.stat().st_mtime) < cutoff_date:
                    file_path.unlink()
                    deleted_count += 1
            except Exception as e:
                logger.warning(f"Failed to delete {file_path}: {str(e)}")
        
        logger.info(f"Cleaned up {deleted_count} old result files")


class DataExporter:
    """Handles export of data in various formats."""
    
    @staticmethod
    def export_to_csv(data: Union[pd.DataFrame, np.ndarray], 
                     filename: str, 
                     feature_names: Optional[List[str]] = None) -> str:
        """Export data to CSV format."""
        if isinstance(data, np.ndarray):
            if feature_names is None:
                feature_names = [f'feature_{i}' for i in range(data.shape[1])]
            df = pd.DataFrame(data, columns=feature_names)
        else:
            df = data
        
        df.to_csv(filename, index=False)
        logger.info(f"Exported data to {filename}")
        return filename
    
    @staticmethod
    def export_metrics_to_excel(metrics_data: List[Dict[str, Any]], 
                               filename: str) -> str:
        """Export metrics data to Excel format."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV")
            return DataExporter.export_metrics_to_csv(metrics_data, filename.replace('.xlsx', '.csv'))
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Performance metrics sheet
        perf_sheet = wb.active
        perf_sheet.title = "Performance Metrics"
        
        # Add headers
        headers = ['Test Name', 'Timestamp', 'Latency (ms)', 'Throughput (pred/s)', 
                  'Accuracy', 'Error Rate']
        for col, header in enumerate(headers, 1):
            perf_sheet.cell(row=1, column=col, value=header)
        
        # Add data
        for row, metrics in enumerate(metrics_data, 2):
            perf_sheet.cell(row=row, column=1, value=metrics.get('test_name', ''))
            perf_sheet.cell(row=row, column=2, value=metrics.get('timestamp', ''))
            
            performance = metrics.get('performance', {})
            perf_sheet.cell(row=row, column=3, value=performance.get('latency_mean_ms', ''))
            perf_sheet.cell(row=row, column=4, value=performance.get('throughput_preds_per_sec', ''))
            perf_sheet.cell(row=row, column=5, value=performance.get('accuracy', ''))
            perf_sheet.cell(row=row, column=6, value=performance.get('error_rate', ''))
        
        # Save workbook
        wb.save(filename)
        logger.info(f"Exported metrics to {filename}")
        return filename
    
    @staticmethod
    def export_metrics_to_csv(metrics_data: List[Dict[str, Any]], 
                             filename: str) -> str:
        """Export metrics data to CSV format."""
        # Flatten metrics data
        flattened_data = []
        
        for metrics in metrics_data:
            row = {
                'test_name': metrics.get('test_name', ''),
                'timestamp': metrics.get('timestamp', ''),
                'n_samples': metrics.get('n_samples', ''),
                'n_features': metrics.get('n_features', '')
            }
            
            # Add performance metrics
            performance = metrics.get('performance', {})
            for key, value in performance.items():
                row[f'perf_{key}'] = value
            
            # Add drift metrics
            drift = metrics.get('drift', {})
            for key, value in drift.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        row[f'drift_{key}_{sub_key}'] = sub_value
                else:
                    row[f'drift_{key}'] = value
            
            flattened_data.append(row)
        
        # Create DataFrame and save
        df = pd.DataFrame(flattened_data)
        df.to_csv(filename, index=False)
        logger.info(f"Exported metrics to {filename}")
        return filename


if __name__ == "__main__":
    # Test the storage modules
    session_manager = SessionManager()
    
    # Create a test session
    session_id = session_manager.create_session("test_session")
    print(f"Created session: {session_id}")
    
    # Add a test run
    test_results = {
        'performance': {
            'latency_mean_ms': 10.5,
            'throughput_preds_per_sec': 1000,
            'accuracy': 0.95
        },
        'drift': {
            'overall_drift_score': 0.2
        }
    }
    
    test_run_id = session_manager.add_test_run(test_results, "test_run_1")
    print(f"Added test run: {test_run_id}")
    
    # Get session summary
    summary = session_manager.get_session_summary()
    print("Session summary:", summary)
    
    # Test results storage
    results_storage = ResultsStorage()
    results_file = results_storage.save_results(test_results)
    print(f"Saved results to: {results_file}")
    
    # Test data export
    sample_data = np.random.random((100, 5))
    csv_file = DataExporter.export_to_csv(sample_data, "test_data.csv")
    print(f"Exported data to: {csv_file}")
